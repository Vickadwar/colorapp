# Copyright (c) 2024, Victor Mandela and contributors
# For license information, please see license.txt

import frappe
import openpyxl
from io import BytesIO
from frappe import _
from datetime import datetime, timedelta
from frappe.utils.file_manager import save_file
from frappe.utils import getdate
import csv
import os

@frappe.whitelist()
def create_painters_meet_attendance(docname):
    # Fetch the Painters Meet Plan document
    meet_plan = frappe.get_doc("Painters Meet Plan", docname)

    # Check if an attendance record already exists
    existing_attendance = frappe.db.exists('Painters Meet Attendance', {'meet_id': meet_plan.name})
    
    if existing_attendance:
        frappe.throw(f"Attendance for this Painters Meet Plan already exists: {existing_attendance}")

    # Prepare the child table data for Painters Meet Attendance
    attendance_list = []
    for plan_item in meet_plan.painters_meet_invite:
        attendance_list.append({
            'member_name': plan_item.colour_academy_member,
            'mobile_number': plan_item.mobile_number
        })

    # Ensure there's at least one invitee to create attendance
    if not attendance_list:
        frappe.throw('No invitees found to create attendance.')

    # Create the attendance document with mapped fields and data
    try:
        attendance_doc = frappe.get_doc({
            'doctype': 'Painters Meet Attendance',
            'meet_id': meet_plan.name,
            'meet_venue': meet_plan.meet_venue,
            'meet_name': meet_plan.meet_name,
            'meet_type': meet_plan.meet_type,
            'sales_representative': meet_plan.sales_representative,
            'town': meet_plan.town,
            'meet_facilitator': meet_plan.meet_facilitator,
            'attendance_list': attendance_list
        })
        attendance_doc.insert()
        frappe.msgprint('Painters Meet Attendance created successfully.')

        # Update the Painters Meet Plan to set the flag that attendance was created
        meet_plan.attendance_created = 1
        meet_plan.save(ignore_permissions=True)

        # Return the new attendance document name to redirect
        return attendance_doc.name

    except Exception as e:
        frappe.throw(f"Error creating Painters Meet Attendance: {str(e)}")


@frappe.whitelist()
def upload_invitees(file_data, plan_name):
    """Upload invitees from an Excel file and add them to the child table"""
    try:
        # Convert the binary string to bytes (if necessary)
        if isinstance(file_data, str):
            file_data = file_data.encode('latin1')

        # Read the Excel file from the binary data
        file_content = BytesIO(file_data)
        workbook = openpyxl.load_workbook(file_content)
        sheet = workbook.active

        # Fetch the plan document
        plan_doc = frappe.get_doc('Painters Meet Plan', plan_name)

        # Get the current user (the uploader)
        current_user = frappe.session.user

        # Loop through each row in the Excel sheet, skipping the header
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Read first_name, second_name, last_name, and mobile_number
            mobile_number, first_name, second_name, last_name = (row + (None,) * 4)[:4]

            # Ensure mobile number is provided
            if not mobile_number:
                frappe.throw(f"Mobile number is missing for the row: {row}")

            # Convert the mobile number to string in case it's an integer
            mobile_number = str(mobile_number)

            # Ensure the mobile number starts with '0' if it's 9 digits long
            if len(mobile_number) == 9 and not mobile_number.startswith('0'):
                mobile_number = '0' + mobile_number

            # Check if the member exists in Colour App Member by mobile_number
            if not frappe.db.exists('Colour App Member', {'mobile_number': mobile_number}):
                # Create new member if they don't exist
                member = frappe.get_doc({
                    'doctype': 'Colour App Member',
                    'mobile_number': mobile_number,
                    'first_name': first_name or '',  # Set empty if None
                    'second_name': second_name or '',  # Set empty if None
                    'last_name': last_name or None  # Last name might be absent
                })
                member.insert()
                member_name = member.name

                # Since this member does not have a tk_membership_number, create a ToDo
                create_missing_tk_todo(member_name, mobile_number, first_name, second_name, current_user)
            else:
                # If the member exists, retrieve their name
                member_name = frappe.get_value('Colour App Member', {'mobile_number': mobile_number}, 'name')

            # Add the member to the child table in Painters Meet Plan
            plan_doc.append('painters_meet_invite', {
                'colour_academy_member': member_name,
                'mobile_number': mobile_number
            })

        # Save the plan document with new invitees
        plan_doc.save()

        frappe.msgprint('Invitees uploaded successfully')
        return True

    except Exception as e:
        frappe.throw(f"Error uploading invitees: {str(e)}")


def create_missing_tk_todo(member_name, mobile_number, first_name, second_name, assigned_to):
    """Create a ToDo for a member missing tk_membership_number"""
    # Set the due date to 7 days from now
    due_date = (datetime.now() + timedelta(days=7)).date()

    # Create the ToDo record
    todo = frappe.get_doc({
        'doctype': 'ToDo',
        'description': f'Please confirm the TK Membership Number for {first_name} {second_name} (Mobile: {mobile_number}) and update it in the system.',
        'status': 'Open',
        'owner': assigned_to,
        'allocated_to': assigned_to,  # Assign task to uploader
        'reference_type': 'Colour App Member',
        'reference_name': member_name,
        'priority': 'Medium',
        'date': due_date
    })
    todo.insert()

    # Trigger Notification to appear under the bell icon
    create_todo_notification(assigned_to, todo)

    frappe.msgprint(f"ToDo created for {first_name} {second_name} to confirm TK Membership Number.")

def create_todo_notification(assigned_to, todo):
    """Send a notification to the user under the notification bell"""
    # Create the notification log entry
    notification_doc = frappe.get_doc({
        'doctype': 'Notification Log',
        'for_user': assigned_to,
        'subject': 'New ToDo Assigned',
        'email_content': f'A new ToDo has been assigned to you: {todo.description}',
        'document_type': 'ToDo',
        'document_name': todo.name,
    })
    notification_doc.insert()

    # Trigger real-time notification refresh (to update the bell count)
    frappe.publish_realtime(
        event='notification',
        user=assigned_to,
        message='New ToDo Assigned'
    )

@frappe.whitelist()
def download_invitees_template():
    """Generate an Excel file template for uploading invitees"""
    # Create an in-memory Excel file
    output = BytesIO()
    
    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Invitees Template'
    
    # Add column headers
    headers = ['Mobile Number', 'First Name', 'Second Name', 'Last Name']
    sheet.append(headers)

    # Save the workbook into the in-memory file
    workbook.save(output)
    output.seek(0)  # Move the cursor to the beginning of the file

    # Send the file to the client
    frappe.response['type'] = 'binary'
    frappe.response['filename'] = 'Invitees_Template.xlsx'
    frappe.response['filecontent'] = output.read()
    frappe.response['content_type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

@frappe.whitelist()
def update_todo_status(doc, method):
    """Automatically close the ToDo when tk_membership_number is updated"""
    if doc.tk_membership_number:
        todos = frappe.get_all("ToDo", filters={
            'reference_type': 'Colour App Member',
            'reference_name': doc.name,
            'status': 'Open'
        })

        # Close all relevant open ToDos
        for todo in todos:
            frappe.db.set_value('ToDo', todo.name, 'status', 'Closed')
            frappe.db.commit()
            frappe.msgprint(f"ToDo for {doc.first_name} {doc.second_name} has been marked as Closed.")


@frappe.whitelist()
def update_dealer_invite_count_on_save(doc, method):
    """Update the invited_counter_staff field for each dealer in the dealers_meet_invite table on save."""
    # Initialize a dictionary to count the number of counter staff for each dealer
    dealer_counts = {}

    # Loop through each counter staff in the counter_staff_meet_invite child table
    for row in doc.counter_staff_meet_invite:
        # Fetch the primary_dealer linked to this counter staff member from the Counter Staff Member Doctype
        primary_dealer = frappe.db.get_value('Counter Staff Member', row.counter_staff_name, 'primary_dealer')

        if primary_dealer:
            # Increment the count for this dealer in dealer_counts
            if primary_dealer in dealer_counts:
                dealer_counts[primary_dealer] += 1
            else:
                dealer_counts[primary_dealer] = 1

    # Now update the invited_counter_staff field in the dealers_meet_invite child table
    for dealer_row in doc.dealers_meet_invite:
        # Match the dealer_name in the dealers_meet_invite table with the primary_dealer
        dealer_row.invited_counter_staff = dealer_counts.get(dealer_row.dealer_name, 0)

    # No need to call save, this is done as part of the save process in the validate event


@frappe.whitelist()
def validate_counter_staff(doc, method=None):
    """Ensure there are no duplicate counter staff and dealers in their respective tables."""
    # Check for duplicate counter staff
    counter_staff_set = set()

    for row in doc.counter_staff_meet_invite:
        if row.counter_staff_name in counter_staff_set:
            frappe.throw(f"Duplicate counter staff {row.counter_staff_name} found. Counter staff must be unique.")
        counter_staff_set.add(row.counter_staff_name)

    # Check for duplicate dealers using dealer_code
    dealer_set = set()

    for dealer_row in doc.dealers_meet_invite:
        dealer_code = frappe.get_value("CPK Dealer", dealer_row.dealer_name, "dealer_code")
        if dealer_code in dealer_set:
            frappe.throw(f"Duplicate dealer with code {dealer_code} found. Dealers must be unique.")
        dealer_set.add(dealer_code)


@frappe.whitelist()
def create_counter_staff_meet_attendance(docname):
    # Fetch the Counter Staff Meet Plan document
    meet_plan = frappe.get_doc("Counter Staff Meet Plan", docname)

    # Check if an attendance record already exists
    existing_attendance = frappe.db.exists('Counter Staff Meet Attendance', {'counter_staff_meet_id': meet_plan.name})
    
    if existing_attendance:
        frappe.throw(f"Attendance for this Counter Staff Meet Plan already exists: {existing_attendance}")

    # Prepare the child table data for Counter Staff Meet Attendance
    dealer_attendance_list = []
    counter_staff_attendance_list = []

    # Copy dealers data
    for dealer in meet_plan.dealers_meet_invite:
        dealer_attendance_list.append({
            'dealer_name': dealer.dealer_name,
            'invited_counter_staff': dealer.invited_counter_staff
        })

    # Copy counter staff data
    for staff in meet_plan.counter_staff_meet_invite:
        counter_staff_attendance_list.append({
            'counter_staff_name': staff.counter_staff_name,
            'mobile_number': staff.mobile_number,
            'dealer': staff.dealer
        })

    # Create the attendance document with mapped fields and data
    try:
        attendance_doc = frappe.get_doc({
            'doctype': 'Counter Staff Meet Attendance',
            'counter_staff_meet_id': meet_plan.name,
            'meet_venue': meet_plan.meet_venue,
            'counter_staff_meet_name': meet_plan.meet_name,
            'meet_type': meet_plan.meet_type,
            'sales_representative': meet_plan.sales_representative,
            'town': meet_plan.town,
            'meet_facilitator': meet_plan.meet_facilitator,
            'dealers': dealer_attendance_list,
            'attendance_list': counter_staff_attendance_list
        })
        attendance_doc.insert()

        # Mark attendance as created in the meet plan
        meet_plan.attendance_created = 1  # Set attendance_created to True/1
        meet_plan.save()

        frappe.msgprint('Counter Staff Meet Attendance created successfully.')

        # Return the new attendance document name to redirect
        return attendance_doc.name

    except Exception as e:
        frappe.throw(f"Error creating Counter Staff Meet Attendance: {str(e)}")


@frappe.whitelist()
def update_execution_status(doc, method):
    """Update the execution status of the related Counter Staff Meet Plan when attendance is submitted."""
    # Check if the attendance document is linked to a Counter Staff Meet Plan
    if doc.counter_staff_meet_id:
        # Fetch the corresponding Counter Staff Meet Plan
        meet_plan = frappe.get_doc("Counter Staff Meet Plan", doc.counter_staff_meet_id)
        
        # Update the execution status to "Executed"
        meet_plan.execution_status = "Executed"
        
        # Save the updated meet plan
        meet_plan.save(ignore_permissions=True)
        # Display a more user-friendly message
        frappe.msgprint(f"The Counter Staff Meet Plan '{meet_plan.meet_name}' has been executed successfully.")


def update_painters_execution_status(doc, method):
    """Update the execution status of the related Painters Meet Plan when attendance is submitted and map skills/products to attendees."""
    # Check if the attendance document is linked to a Painters Meet Plan
    if doc.meet_id:
        # Fetch the corresponding Painters Meet Plan
        meet_plan = frappe.get_doc("Painters Meet Plan", doc.meet_id)

        # Update the execution status to "Executed"
        meet_plan.execution_status = "Executed"
        meet_plan.save(ignore_permissions=True)

        # Ensure the child tables are properly populated before proceeding
        if not doc.skills or not doc.products:
            frappe.throw("Skills or Products tables are empty. Please add training details before submitting.")

        # Map skills and products to each Colour App Member
        map_skills_and_products_to_members(doc)

        frappe.msgprint(f"The Painters Meet Plan '{meet_plan.meet_name}' has been executed successfully.")

def map_skills_and_products_to_members(doc):
    """Map the training skills and products to each attendee's Colour App Member profile."""
    # Use the meet_date from the attendance document
    training_date = doc.meet_date

    # Iterate over the list of attendees
    for attendee in doc.attendance_list:
        member_name = attendee.member_name
        member = frappe.get_doc('Colour App Member', member_name)

        # Ensure attendee's profile exists
        if not member:
            frappe.throw(f"Member with name '{member_name}' not found.")

        # Map skills from the attendance to the member's profile in 'skills_trained_on'
        for skill in doc.skills:
            member.append('skills_trained_on', {
                'skill_name': skill.skill_name,
                'training_date': training_date  # Use the meet_date from the attendance
            })
        
        # Map products from the attendance to the member's profile in 'products_trained_on'
        for product in doc.products:
            member.append('products_trained_on', {
                'product_name': product.product_name,
                'training_date': training_date  # Use the meet_date from the attendance
            })
        
        # Update the last_training_date field with the meet_date from the attendance
        member.last_training_date = training_date

        # Save the updated Colour App Member with mapped skills, products, and the last_training_date
        member.save(ignore_permissions=True)

        # Log successful mapping
        frappe.msgprint(f"Skills, products, and last training date for {member_name} have been updated successfully.")

@frappe.whitelist()
def get_item_balance(item_code, warehouse):
    """
    Fetch the latest stock balance from the Merchandise Ledger for a given item and warehouse.
    This will be called via a client-side script.
    """
    last_entry = frappe.db.get_value(
        "Merchandise Ledger",
        {"merchandise_item_code": item_code, "merchandise_warehouse": warehouse},
        "balance_after",
        order_by="posting_datetime desc"
    )
    return last_entry or 0  # Return 0 if no previous stock entry is found

#Creates a download button on the Painters Meet Attendance Doctype, where the document has succefully been submitted.
@frappe.whitelist()
def download_attendance_excel(docname):
    # Fetch the Painters Meet Attendance document
    doc = frappe.get_doc('Painters Meet Attendance', docname)

    # Create a new Excel file in memory
    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Attendance List'

    # Fetch the actual venue name from the Meet Venue Doctype using the correct field
    meet_venue_name = frappe.db.get_value('Meet Venue', doc.meet_venue, 'meet_venue_name')
    # Fallback to venue ID if the venue name is not found
    meet_venue_name = meet_venue_name if meet_venue_name else doc.meet_venue

    # Add headers for the meet details with the correct venue name
    sheet.append(['Meet Name:', doc.meet_name])
    sheet.append(['Meet ID:', doc.meet_id])
    sheet.append(['Meet Facilitator:', doc.meet_facilitator])
    sheet.append(['Meet Venue:', meet_venue_name])  # Use the resolved venue name
    sheet.append(['Meet Type:', doc.meet_type])
    sheet.append(['Meet Date:', str(doc.meet_date)])
    
    # Add a blank row between meet details and attendance list
    sheet.append([''])

    # Add headers for the attendee list
    sheet.append(['Member Name', 'Mobile Number'])

    # Fetch the attendee list from the child table (Painters Meet Attendance Detail)
    for attendee in doc.attendance_list:
        # Fetch the full name from the Colour App Member Doctype
        member_full_name = frappe.db.get_value('Colour App Member', attendee.member_name, 'member_name')
        member_full_name = member_full_name if member_full_name else attendee.member_name

        # Append the full name and mobile number to the sheet
        sheet.append([member_full_name, attendee.mobile_number])

    # Save the workbook to the output stream
    workbook.save(output)
    output.seek(0)

    # Create a file and store it in the file system
    file_name = f"Attendance_List_{docname}.xlsx"
    saved_file = save_file(file_name, output.read(), doc.doctype, docname, is_private=1)

    # Return the file URL to the frontend
    return saved_file.file_url

#Creates a doenload button for the Counter Staff Attendnace DOctype for successfully submitted Document. 
@frappe.whitelist()
def download_counter_staff_meet_attendance_excel(docname):
    # Fetch the Counter Staff Meet Attendance document
    doc = frappe.get_doc('Counter Staff Meet Attendance', docname)

    # Create a new Excel file in memory
    output = BytesIO()
    workbook = openpyxl.Workbook()

    # Create a worksheet for the parent Doctype and meet details
    sheet = workbook.active
    sheet.title = 'Counter Staff Meet Attendance'

    # Fetch the venue name from the linked Meet Venue Doctype
    meet_venue_name = frappe.db.get_value('Meet Venue', doc.meet_venue, 'meet_venue_name')
    meet_venue_name = meet_venue_name if meet_venue_name else doc.meet_venue

    # Add headers for the Counter Staff Meet Attendance details
    sheet.append(['Counter Staff Meet Name:', doc.counter_staff_meet_name])
    sheet.append(['Meet Date:', str(doc.meet_date)])
    sheet.append(['Meet Venue:', meet_venue_name])
    sheet.append(['Meet Facilitator:', doc.meet_facilitator])
    sheet.append(['Counter Staff Meet ID:', doc.counter_staff_meet_id])

    # Add a blank row between the meet details and the child table data
    sheet.append([''])

    # Add headers for the first child table: Counter Staff Dealer Attendance Detail
    sheet.append(['Dealers'])
    sheet.append(['Dealer Name', 'Invited Counter Staff'])

    # Fetch the dealer data from the first child table
    for dealer in doc.dealers:
        # Fetch the dealer's full name from the CPK Dealer Doctype
        dealer_name = frappe.db.get_value('CPK Dealer', dealer.dealer_name, 'dealer_name')
        dealer_name = dealer_name if dealer_name else dealer.dealer_name

        # Append the dealer data to the sheet
        sheet.append([dealer_name, dealer.invited_counter_staff])

    # Add a blank row between the dealer data and the attendance data
    sheet.append([''])

    # Add headers for the second child table: Counter Staff Meet Invite Attendance Detail
    sheet.append(['Attendance'])
    sheet.append(['Counter Staff Name', 'Mobile Number'])

    # Fetch the attendance data from the second child table
    for attendee in doc.attendance_list:
        # Fetch the full name from the Counter Staff Member Doctype
        counter_staff_name = frappe.db.get_value('Counter Staff Member', attendee.counter_staff_name, 'counter_staff_name')
        counter_staff_name = counter_staff_name if counter_staff_name else attendee.counter_staff_name

        # Append the counter staff attendance data to the sheet
        sheet.append([counter_staff_name, attendee.mobile_number])

    # Save the workbook to the output stream
    workbook.save(output)
    output.seek(0)

    # Create a file and store it in the file system
    file_name = f"Counter_Staff_Meet_Attendance_List_{docname}.xlsx"
    saved_file = save_file(file_name, output.read(), doc.doctype, docname, is_private=1)

    # Return the file URL to the frontend
    return saved_file.file_url

#Meet Download Dump goes here ------------------- CHECK CHECK
@frappe.whitelist()
def download_single_meet_painters(meet_name, docname):
    attendance_data = frappe.get_doc("Painters Meet Attendance", meet_name)
    if attendance_data.docstatus != 1:  # Only download submitted documents
        frappe.throw(_("Selected meet has not been submitted."))

    # Prepare the data
    rows = prepare_single_meet_data(attendance_data, 'painters')
    metadata = {
        "meet_name": attendance_data.meet_name,
        "meet_date": attendance_data.get('meet_date'),
        "meet_venue": frappe.get_value("Meet Venue", attendance_data.meet_venue, "name"),
        "meet_facilitator": attendance_data.get('meet_facilitator'),
        "meet_id": attendance_data.get('meet_id'),
    }
    create_custom_csv_with_layout(rows, "single_meet_painters.csv", docname, metadata)

@frappe.whitelist()
def download_single_meet_counter_staff(meet_name, docname):
    attendance_data = frappe.get_doc("Counter Staff Meet Attendance", meet_name)
    if attendance_data.docstatus != 1:  # Only download submitted documents
        frappe.throw(_("Selected meet has not been submitted."))

    # Prepare the data
    rows = prepare_single_meet_data(attendance_data, 'counter_staff')
    metadata = {
        "meet_name": attendance_data.counter_staff_meet_name,
        "meet_date": attendance_data.get('meet_date'),
        "meet_venue": frappe.get_value("Meet Venue", attendance_data.meet_venue, "name"),
        "meet_facilitator": attendance_data.get('meet_facilitator'),
        "meet_id": attendance_data.get('counter_staff_meet_id'),
    }
    create_custom_csv_with_layout(rows, "single_meet_counter_staff.csv", docname, metadata)

def prepare_single_meet_data(attendance_data, meet_type):
    # Conditional logic based on meet type to handle different field names
    data = {
        "meet_name": attendance_data.meet_name if meet_type == 'painters' else attendance_data.counter_staff_meet_name,
        "meet_id": attendance_data.get('meet_id') if meet_type == 'painters' else attendance_data.get('counter_staff_meet_id'),
        "meet_venue": frappe.get_value("Meet Venue", attendance_data.meet_venue, "name"),
        "meet_date": attendance_data.get('meet_date'),
        "meet_facilitator": attendance_data.get('meet_facilitator'),
    }

    # Prepare rows for each participant
    rows = []
    for participant in attendance_data.attendance_list:
        row = data.copy()
        row.update({
            "member_name": frappe.get_value("Colour App Member", participant.member_name, "name") if meet_type == 'painters' else frappe.get_value("Counter Staff Member", participant.counter_staff_name, "name"),
            "mobile_number": participant.mobile_number,
            "gift_one": participant.gift_one,
            "gift_two": participant.gift_two
        })
        rows.append(row)
    
    return rows

@frappe.whitelist()
def download_multiple_meets(meet_type, from_date, to_date, docname):
    # Fetch the correct doctype and field names based on the meet_type
    if meet_type == "Painters Meet":
        doctype = "Painters Meet Attendance"
        meet_name_field = "meet_name"
        meet_id_field = "meet_id"
    else:  # Counter Staff Meet
        doctype = "Counter Staff Meet Attendance"
        meet_name_field = "counter_staff_meet_name"
        meet_id_field = "counter_staff_meet_id"
    
    # Retrieve all meets within the date range and ensure the correct fields are fetched
    meets = frappe.get_all(doctype,
                           filters={
                               "meet_date": [">=", getdate(from_date)],
                               "meet_date": ["<=", getdate(to_date)],
                               "docstatus": 1  # Only submitted documents
                           },
                           fields=["name", meet_name_field, meet_id_field, "meet_venue", "meet_date", "meet_facilitator"])

    rows = []
    for meet in meets:
        attendance_data = frappe.get_doc(doctype, meet['name'])  # Correct access to meet data
        meet_data = {
            "meet_name": meet[meet_name_field],
            "meet_id": meet[meet_id_field],
            "meet_venue": frappe.get_value("Meet Venue", attendance_data.meet_venue, "name"),
            "meet_date": attendance_data.get('meet_date'),
            "meet_facilitator": attendance_data.get('meet_facilitator'),
        }
        
        for participant in attendance_data.attendance_list:
            row = meet_data.copy()
            row.update({
                "member_name": frappe.get_value("Colour App Member", participant.member_name, "name") if meet_type == "Painters Meet" else frappe.get_value("Counter Staff Member", participant.counter_staff_name, "name"),
                "mobile_number": participant.mobile_number,
                "gift_one": participant.gift_one,
                "gift_two": participant.gift_two
            })
            rows.append(row)

    create_custom_csv_with_layout(rows, "multiple_meets.csv", docname, {})  # Generate CSV

@frappe.whitelist()
def download_painters_dump(from_date, to_date, docname):
    # First, get the meets that occurred within the date range
    meets = frappe.get_all(
        "Painters Meet Attendance",
        filters={
            "meet_date": [">=", getdate(from_date)],
            "meet_date": ["<=", getdate(to_date)],
            "docstatus": 1  # Only submitted events
        },
        fields=["name"]
    )

    # Get the members who attended these meets
    members = []
    for meet in meets:
        attendance_data = frappe.get_doc("Painters Meet Attendance", meet['name'])
        for participant in attendance_data.attendance_list:
            member_data = frappe.get_value("Colour App Member", participant.member_name, ["name", "mobile_number", "last_training_date"], as_dict=True)
            # Ensure we add only if this member doesn't have a duplicate entry with the same last_training_date
            if not any(m['name'] == member_data['name'] and m['last_training_date'] == member_data['last_training_date'] for m in members):
                members.append(member_data)

    # Create CSV with just the necessary columns
    create_custom_csv_with_layout(members, "painters_dump.csv", docname, {})

@frappe.whitelist()
def download_counter_staff_dump(from_date, to_date, docname):
    # First, get the meets that occurred within the date range
    meets = frappe.get_all(
        "Counter Staff Meet Attendance",
        filters={
            "meet_date": [">=", getdate(from_date)],
            "meet_date": ["<=", getdate(to_date)],
            "docstatus": 1  # Only submitted events
        },
        fields=["name"]
    )

    # Get the members who attended these meets
    members = []
    for meet in meets:
        attendance_data = frappe.get_doc("Counter Staff Meet Attendance", meet['name'])
        for participant in attendance_data.attendance_list:
            member_data = frappe.get_value("Counter Staff Member", participant.counter_staff_name, ["name", "mobile_number", "last_training_date"], as_dict=True)
            # Ensure we add only if this member doesn't have a duplicate entry with the same last_training_date
            if not any(m['name'] == member_data['name'] and m['last_training_date'] == member_data['last_training_date'] for m in members):
                members.append(member_data)

    # Create CSV with just the necessary columns
    create_custom_csv_with_layout(members, "counter_staff_dump.csv", docname, {})

# Adjusted CSV creation function
def create_custom_csv_with_layout(data, filename, docname, metadata):
    # Ensure the private/files directory exists
    directory = frappe.get_site_path('private', 'files')
    if not os.path.exists(directory):
        os.makedirs(directory)

    # Generate the full file path
    file_path = os.path.join(directory, filename)
    
    with open(file_path, mode='w', newline='') as file:
        writer = csv.writer(file)

        # Write the table headers
        headers = ["Member Name", "Mobile Number", "Last Training Date"]
        writer.writerow(headers)

        # Write participant rows with proper spacing
        for row in data:
            writer.writerow([
                row.get("name", ""),
                row.get("mobile_number", ""),
                row.get("last_training_date", "")
            ])
    
    # Generate the file URL for attaching the document
    file_url = os.path.join('/private/files', filename)
    
    # Attach the file to the current document
    attach_file_to_document(docname, filename, file_url)

def attach_file_to_document(docname, filename, file_url):
    file_doc = frappe.get_doc({
        'doctype': 'File',
        'file_name': filename,
        'attached_to_doctype': 'Meet Dump Download',
        'attached_to_name': docname,
        'file_url': file_url,
        'is_private': 1,  # Set to private
    })
    file_doc.save()